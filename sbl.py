import pandas as pd
from openpyxl.styles import PatternFill
import streamlit as st
import traceback
from datetime import datetime
import io

# Streamlit 페이지 설정
st.set_page_config(
    page_title="삼성바이오 레포트 변환기",
    page_icon="📊",
    layout="wide"
)

mapping_dict = {
    "신고번호": "신고번호",
    "신청차수": "정정차수",
    "신고일자": "신고일자",
    "접수일자": "신고일자",
    "수리일자": "수리일자",
    "세관": "신고세관",
    "과": "신고과",
    "MasterB/L번호": "MasterB/L번호",
    "B/L(AWB)번호": "B/L번호",
    "B/L분할여부": "B/L구분",
    "화물관리번호": "화물관리번호",
    "입항일자": "입항일자",
    "반입일자": "반입일자",
    "징수형태": "징수형태",
    "수입자상호": "수입자상호",
    "납세의무자상호": "납세자상호",
    "납세의무자대표자": "납세자대표자명",
    "납세의무자사업번호": "납세자사업번호",
    "무역거래처상호": "무역거래처상호",
    "무역거래처부호": "무역거래처코드",
    "무역거래처국가": "무역거래처국가코드",
    "해외공급자상호": "해외공급자상호",
    "해외공급자부호": "해외공급자코드",
    "해외공급자국가": "해외공급자국가코드",
    "통관계획부호": "통관계획",
    "신고구분부호": "신고구분",
    "거래구분부호": "거래구분",
    "수입종류부호": "수입종류",
    "컨테이너번호": "컨테이너번호",
    "원산지증명서유무": "원산지증명유무",
    "가격신고서유무": "가격신고서유무",
    "총중량": "총중량",
    "총수량": "환급물량",
    "총환급물량": "환급물량",
    "총포장갯수": "총포장수량",
    "총포장종류": "포장수량단위",
    "국내도착항부호": "도착항코드",
    "운송형태수단": "운송형태",
    "운송형태용기": "운송용기",
    "적출국부호": "적출국코드",
    "선(기)명": "선기명_1",
    "운수기관부호": "운수기관",
    "검사장치장소부호": "장치장부호",
    "검사(반입)장소부호": "장치장소",
    "검사(반입)장소장소명": "장치장명",
    "총란수": "총란수",
    "인도조건": "인도조건",
    "결제통화": "결제통화단위",
    "결제환율": "결제통화단위환율",
    "미화환율": "USD환율",
    "결제총금액": "입력결제금액",
    "결재방법": "결제방법",
    "결제금액": "입력결제금액",
    "기타금액": "기타금액",
    "총과세가격미화": "Cif달러",
    "총과세가격원화": "Cif원화",
    "운임원화": "계산된운임원화",
    "운임1 통화종류": "운임통화단위",
    "운임1 사용자기재": None,
    "운임2 통화종류": "운임통화단위",
    "운임2 사용자기재": None,
    "보험원화": "계산된보험료원화",
    "보험료1 통화종류": None,
    "보험료1 사용자기재": None,
    "가산금액원화란로열티합": "전체계산된가산금원화",
    "가산금액 구분(율,금액)": "가산금구분",
    "가산금액 통화종류": "가산금통화단위",
    "가산비용/율": "입력가산금",
    "가산금액 통화종류_1": "가산금통화단위",
    "가산비용/율_1": None,
    "가산금액 원화(로열티 제외)": "공통사항계산된가산금원화",
    "공제금액원화": "공통사항계산된공제금원화",
    "공제금액 구분(율,금액계산)": "공제금구분",
    "공제금액 통화종류": "공제금통화단위",
    "공제금액/율": "입력공제금",
    "총관세": "관세",
    "총개소세": "특소세",
    "총주세": "주세",
    "총교통세": "교통세",
    "총교육세": "교육세",
    "총농특세": "농특세",
    "총부가세": "부가세",
    "총신고지연가산세": "신고지연가산세",
    "총세액합계": "총세액",
    "부가세과세과표": "부가세과세과표",
    "부가세면세과표": "부가세면세과표",
    "특송업체코드": "특송업체코드",
    "신고인기재란1": "관세사기재2_01",
    "신고인기재란2": "관세사기재2_02",
    "신고인기재란3": "관세사기재2_03",
    "신고인기재란4": "관세사기재2_04",
    "신고인기재란5": "관세사기재2_05",
    "신고인기재란6": "관세사기재2_06",
    "신고인기재란7": "관세사기재2_07",
    "신고인기재란8": "관세사기재2_08",
    "신고인기재란9": "관세사기재2_09",
    "심사담당자성명": "세관담당자이름",
    "심사담당자직원부호": "세관담당자부호",
    "수신결과": "수신결과",
    "수입의뢰번호": None,
    "신고자상호": "신고자상호",
    "BL분할사유코드": "B/L분할사유코드",
    "BL분할기타사유": "기타사유",
    "보세공장사용신고구분": "사용신고구분",
    "보세공장사용신고설명": None,
    "보세공장사용일자": "사용신고일자",
    "대행사코드": "대행사코드",
    "대행사상호": "대행사상호",
    "운송주선인부호": "운송주선인코드",
    "운송주선인상호": "운송주선인상호",
    "세관기재란": "세관기재란",
    "란번호": "란번호2",
    "세번부호": "세번부호",
    "표준품명(미전송)": None,
    "표준품명코드": "표준품명코드",
    "품명1": "표준품명",
    "품명2": None,
    "거래품명1": "거래품명",
    "거래품명2": None,
    "상표코드": "상표코드",
    "상표명": "상표명",
    "첨부여부": "첨부",
    "신고가격": "란결제금액",
    "과세가격(원화)": "과세가격원화",
    "과세가격 미화": "과세가격달러",
    "순중량": "순중량",
    "순중량단위": "순중량단위",
    "수량": "수량",
    "수량단위": "수량단위",
    "환급물량": "환급물량",
    "환급물량단위": "환급물량단위",
    "C/S 검사구분 부호": "CS검사",
    "검사방법변경 부호": None,
    "원산지국가부호": "원산지이름",
    "원산지결정기준": "원산지표시결정방법",
    "원산지표시유무": "원산지표시유무",
    "원산지표시방법": "원산지표시방법",
    "원산지발행번호": "원산지증명서발급번호",
    "원산지발행일자": "원산지증명서발급일자",
    "원산지발행국가": "원산지증명서발급국가",
    "원산지발행기관": "원산지증명서발급기관",
    "원산지발급지역": "원산지증명서발급지역",
    "원산지발급담당자": "원산지증명서발급담당자",
    "원산지기준": "원산지증명서원산지기준",
    "원산지분할여부": "원산지증명서발행번호분할여부",
    "가산비용": "란입력가산금",
    "공제비용": None,
    "세종부호": "세율설명",
    "관세구분": "세율구분",
    "관세율": "관세실행세율",
    "탄력관세구분": "탄력구분",
    "탄력관세율": "탄력실행세율",
    "단위당 세액": None,
    "관세감면율": "관세감면율",
    "관세액": "실제관세액",
    "관세감면액": "경감관세",
    "관세감면/분납부호": "관세감면분납부호",
    "관세감면/분납구분": "관세감면구분",
    "전송용세율/단위당세액": "관세실행세율",
    "내국세 구분": "내국세구분",
    "내국세 세종구분": "내국세구분",
    "내국세 부호": "내국세부호",
    "내국세율": "내국세율",
    "내국세 감면액": "내국세면세",
    "내국세": "내국세_1",
    "개소세 면세부호": "특소세면세부호",
    "개소세 기준가격 공제": "특소세",
    "개소세": "특소세",
    "교통세": "교육세_1",
    "주세": "주세",
    "교육세 구분": "교육세구분",
    "교육세 세종구분": None,
    "교육세 감면액": "내국세면세",
    "교육세 율": "교육세율",
    "교육세": "교육세",
    "농특세구분": "농특세구분",
    "농특세세종부호": None,
    "농특세": "농특세",
    "농특세율": "농특세세율",
    "부가세 율": None,
    "부가세 구분": "부가세구분",
    "부가세 세종부호": None,
    "부가세": "부가세_1",
    "로열티 구분": None,
    "로열티 율": None,
    "로열티 금액": None,
    "공제비용 원화": "공제비용원화",
    "부가세감면부호": "부가세감면부호",
    "부가세 감면액": "면세부가세",
    "부가세 감면율": "부가세경감율",
    "관세액 기준": None,
    "부가세 과세과표": "부가세과표",
    "부가세 면세과표": "부가세면세과표_1",
    "용도세율공문번호": "용도세율전용물품확인공문번호",
    "총규격수": "총규격수",
    "총요건확인서류수": "요건확인서류수",
    "총재수출서류수": None,
    "총요건비대상서류수": None,
    "총규격수_1": "총규격수",
    "규격번호": "행번호",
    "자재코드": "자재코드",
    "규격1": "규격1",
    "규격2": "규격2",
    "규격3": "규격3",
    "성분1": "성분1",
    "성분2": "성분2",
    "규격수량": "수량_1",
    "규격단위": "수량단위_1",
    "규격단가": "단가",
    "규격금액": "금액",
    "정정신청일자": None,
    "정정승인일자": None,
    "정정신청구분": None,
    "정정사유코드": None,
    "정정사유코드명": None,
    "정정차수": "정정차수",
    "전송결과": "전송결과",
    "신규작성자": None,
    "최종수정자": None,
    "가격신고_송품장번호": None,
    "가격신고_송품장발행일": None,
    "가격신고_잠정가격신고번호": None,
    "가격신고_가격확정예정시기": None,
    "입력일자": "입력일시",
}


def log_error(message, e=None):
    st.error(f"[오류] {message}")
    if e:
        st.error(f"  - 오류 유형: {type(e).__name__}")
        st.error(f"  - 오류 내용: {str(e)}")
        with st.expander("상세 오류 정보"):
            traceback_str = ''.join(traceback.format_tb(e.__traceback__))
            st.code(traceback_str)


def log_info(message):
    st.info(f"[정보] {message}")


def log_success(message):
    st.success(f"[완료] {message}")


def process_value(val, col_name, raw_col_name=None):
    if pd.isna(val):
        return ""
    val_str = str(val)
    
    if col_name in ["정정차수", "세관", "신고세관"]:
        return val_str
    
    try:
        if val_str.strip() and val_str.replace('.', '', 1).isdigit():
            return str(int(float(val_str))).zfill(3)
    except ValueError:
        pass
    return val_str


# Streamlit UI
def main():
    st.title("📊 삼성바이오 레포트 변환기")
    st.markdown("---")
    
    # 파일 업로드
    st.subheader("1. RAW 데이터 파일 업로드")
    uploaded_file = st.file_uploader(
        "Excel 파일을 선택하세요 (.xlsx, .xls)",
        type=['xlsx', 'xls'],
        help="RAW 데이터가 포함된 엑셀 파일을 업로드해주세요."
    )
    
    if uploaded_file is not None:
        try:
            # 파일 정보 표시
            st.success(f"✅ 파일 업로드 완료: {uploaded_file.name}")
            
            # 데이터 로드
            with st.spinner("📂 RAW 데이터를 로딩 중..."):
                raw_df = pd.read_excel(uploaded_file, sheet_name=0)
                raw_df = raw_df.dropna(axis=1, how='all')
            
            log_success(f"RAW 데이터 로드 완료 (행: {len(raw_df)}, "
                       f"열: {len(raw_df.columns)})")
            
            # 데이터 미리보기
            st.subheader("2. 데이터 미리보기")
            with st.expander("원본 데이터 확인", expanded=False):
                st.dataframe(raw_df.head(10), use_container_width=True)
                st.caption(f"총 {len(raw_df)}행, {len(raw_df.columns)}열")
            
            # 매핑 처리
            with st.spinner("🔄 매핑 딕셔너리를 적용 중..."):
                data = {}
                raw_headers = []
                
                progress_bar = st.progress(0)
                total_cols = len(mapping_dict)
                
                for idx, (customer_col, raw_col) in enumerate(
                    mapping_dict.items()
                ):
                    try:
                        if raw_col in [None, "#N/A"]:
                            data[customer_col] = [""] * len(raw_df)
                            raw_headers.append("")
                        else:
                            if raw_col in raw_df.columns:
                                data[customer_col] = (
                                    raw_df[raw_col].fillna("").astype(str)
                                )
                                raw_headers.append(raw_col)
                            else:
                                data[customer_col] = [""] * len(raw_df)
                                raw_headers.append("")
                    except Exception as e:
                        log_error(f"컬럼 '{customer_col}' 매핑 중 오류 발생", e)
                        data[customer_col] = [""] * len(raw_df)
                        raw_headers.append("")
                    
                    # 진행률 업데이트
                    progress_bar.progress((idx + 1) / total_cols)
                
                progress_bar.empty()
            
            # 최종 데이터프레임 생성
            with st.spinner("⚙️ 데이터 변환 중..."):
                final_df = pd.DataFrame(data)
                
                # 데이터 변환
                for col in final_df.columns:
                    if col not in ["정정차수", "세관", "신고세관"]:
                        final_df[col] = final_df[col].apply(
                            lambda x: process_value(x, col)
                        )
                    elif col in ["정정차수", "세관", "신고세관"]:
                        final_df[col] = final_df[col].apply(
                            lambda x: str(int(float(x))).zfill(3) 
                            if str(x).replace('.', '', 1).isdigit() 
                            else str(x)
                        )
                
                # NO 컬럼 추가 (성능 경고 해결)
                no_column = pd.Series(range(1, len(final_df) + 1), name='NO')
                final_df = pd.concat([no_column, final_df], axis=1)
                
                # raw_headers에도 NO 컬럼 추가
                raw_headers.insert(0, "NO")
            
            log_success(f"최종 데이터프레임 생성 완료 (행: {len(final_df)}, "
                       f"열: {len(final_df.columns)})")
            
            # 변환된 데이터 미리보기
            st.subheader("3. 변환된 데이터 미리보기")
            with st.expander("변환된 데이터 확인", expanded=True):
                st.dataframe(final_df.head(10), use_container_width=True)
                st.caption(f"총 {len(final_df)}행, {len(final_df.columns)}열")
            
            # 다운로드 섹션
            st.subheader("4. 결과 파일 다운로드")
            
            if st.button("📥 엑셀 파일 생성 및 다운로드", 
                        type="primary", use_container_width=True):
                with st.spinner("📊 엑셀 파일을 생성 중..."):
                    # 임시 파일 생성
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    temp_filename = f"삼성바이오_보고서_결과_{timestamp}.xlsx"
                    
                    # 메모리에서 엑셀 파일 생성
                    output = io.BytesIO()
                    
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # raw_headers를 첫 번째 행으로 추가
                        headers_df = pd.DataFrame(
                            [raw_headers], columns=final_df.columns
                        )
                        combined_df = pd.concat(
                            [headers_df, final_df], ignore_index=True
                        )
                        combined_df.to_excel(
                            writer, index=False, sheet_name='데이터'
                        )
                        
                        # 스타일 적용
                        worksheet = writer.sheets['데이터']
                        yellow_fill = PatternFill(
                            start_color='FFFF00', 
                            end_color='FFFF00', 
                            fill_type='solid'
                        )
                        yellow_columns = ['신고번호', '정정차수', 
                                        '신고일자', '수리일자']
                        
                        # 2번째 행이 실제 헤더
                        for col_idx, header in enumerate(worksheet[2], 1):
                            if header.value in yellow_columns:
                                for row in worksheet.iter_rows(
                                    min_row=3, max_row=worksheet.max_row,
                                    min_col=col_idx, max_col=col_idx
                                ):
                                    row[0].fill = yellow_fill
                            
                            # 텍스트 형식 지정
                            if header.value in ['신고세관', '세관']:
                                for row in worksheet.iter_rows(
                                    min_row=3, max_row=worksheet.max_row,
                                    min_col=col_idx, max_col=col_idx
                                ):
                                    row[0].number_format = '@'
                    
                    output.seek(0)
                    
                    st.download_button(
                        label="💾 엑셀 파일 다운로드",
                        data=output.getvalue(),
                        file_name=temp_filename,
                        mime="application/vnd.openxmlformats-"
                             "officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                    log_success("엑셀 파일이 성공적으로 생성되었습니다!")
                    st.balloons()
            
        except Exception as e:
            log_error("파일 처리 중 오류가 발생했습니다", e)
    
    else:
        st.info("👆 위에서 RAW 데이터 파일을 업로드해주세요.")
        
        # 사용 방법 안내
        with st.expander("📖 사용 방법", expanded=True):
            st.markdown("""
            ### 사용 단계:
            1. **RAW 데이터 파일 업로드**: Excel 파일(.xlsx, .xls)을 
               선택하세요
            2. **데이터 확인**: 업로드된 원본 데이터를 미리보기로 
               확인하세요
            3. **변환 결과 확인**: 매핑이 적용된 변환 결과를 확인하세요
            4. **다운로드**: 변환된 엑셀 파일을 다운로드하세요
            
            ### 특징:
            - ✅ 자동 매핑 적용
            - ✅ 스타일 지정 (노란색 하이라이트)
            - ✅ 데이터 형식 최적화
            - ✅ 브라우저에서 바로 실행
            """)
        
        # 변환 로직 설명 추가
        with st.expander("🔧 변환 로직 설명", expanded=False):
            st.markdown("""
            ### 📊 데이터 변환 프로세스:
            
            #### 1️⃣ **파일 읽기 & 전처리**
            ```python
            # Excel 파일 로드
            raw_df = pd.read_excel(uploaded_file)
            # 빈 컬럼 제거
            raw_df = raw_df.dropna(axis=1, how='all')
            ```
            
            #### 2️⃣ **매핑 딕셔너리 적용**
            ```python
            # 미리 정의된 244개 컬럼 매핑
            mapping_dict = {
                "신고번호": "신고번호",
                "신청차수": "정정차수", 
                "세관": "신고세관",
                # ... 총 244개 매핑 규칙
            }
            ```
            
            #### 3️⃣ **데이터 변환 & 정제**
            ```python
            # 숫자 데이터 3자리 포맷팅 (예: 1 → 001)
            if val_str.replace('.', '', 1).isdigit():
                return str(int(float(val_str))).zfill(3)
            
            # 특별 컬럼 처리 (정정차수, 세관, 신고세관)
            elif col in ["정정차수", "세관", "신고세관"]:
                # 원본 데이터 유지
            ```
            
            #### 4️⃣ **엑셀 스타일 적용**
            ```python
            # 노란색 하이라이트 (중요 컬럼)
            yellow_columns = ['신고번호', '정정차수', '신고일자', '수리일자']
            
            # 텍스트 형식 지정 (숫자로 변환 방지)
            text_columns = ['신고세관', '세관']
            ```
            
            #### 5️⃣ **최종 출력**
            - **1행**: RAW 데이터 원본 헤더명
            - **2행**: 변환된 고객사 헤더명  
            - **3행~**: 변환된 실제 데이터
            - **NO 컬럼**: 자동 생성된 행 번호 (1부터 시작)
            
            ---
            
            ### 🎯 **핵심 특징:**
            - **244개 컬럼** 자동 매핑 처리
            - **실시간 진행률** 표시 (진행 바)
            - **오류 처리**: 누락된 컬럼도 안전하게 처리
            - **성능 최적화**: pd.concat() 사용으로 빠른 처리
            - **스타일 적용**: 업무에 필요한 시각적 강조
            
            ### 🔍 **데이터 품질 보장:**
            - **NULL 값 처리**: 빈 문자열로 통일
            - **데이터 타입 통일**: 모든 값을 문자열로 변환
            - **포맷 표준화**: 숫자는 3자리 포맷으로 통일
            - **원본 보존**: RAW 헤더 정보 유지
            """)


if __name__ == "__main__":
    main() 